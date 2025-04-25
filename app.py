# from flask import Flask, render_template, request, jsonify
# import cv2
# import numpy as np
# import os
# from datetime import datetime
# import face_recognition
# from openpyxl import Workbook, load_workbook
# from werkzeug.utils import secure_filename

# app = Flask(__name__)

# # Configuration
# STUDENT_IMAGES_PATH = 'StudentImages'
# ROLL_CSV_FILE = 'attendancenew.csv'
# ATTENDANCE_EXCEL_FILE = 'attendancenew.xlsx'

# # Initialize Excel file
# def init_attendance_excel():
#     if os.path.exists(ATTENDANCE_EXCEL_FILE):
#         workbook = load_workbook(ATTENDANCE_EXCEL_FILE)
#         sheet = workbook.active
        
#         if sheet.cell(row=1, column=1).value != "Roll No.":
#             sheet.cell(row=1, column=1, value="Roll No.")
#             sheet.cell(row=1, column=2, value="Name")
#     else:
#         workbook = Workbook()
#         sheet = workbook.active
#         sheet.cell(row=1, column=1, value="Roll No.")
#         sheet.cell(row=1, column=2, value="Name")
    
#     return workbook, sheet

# attendance_workbook, attendance_sheet = init_attendance_excel()
# current_date = datetime.now().strftime('%Y-%m-%d')

# def add_date_columns_if_needed():
#     global attendance_sheet, current_date
    
#     date_col = 3
#     date_exists = False
    
#     while True:
#         header = attendance_sheet.cell(row=1, column=date_col).value
#         if header is None:
#             break
#         if header == current_date:
#             date_exists = True
#             break
#         date_col += 4
    
#     if not date_exists:
#         attendance_sheet.cell(row=1, column=date_col, value=current_date)
#         attendance_sheet.cell(row=2, column=date_col, value="Status")
#         attendance_sheet.cell(row=2, column=date_col+1, value="Timestamp")
    
#     return date_col

# date_column = add_date_columns_if_needed()

# def find_student_row(name):
#     for row in range(3, attendance_sheet.max_row + 1):
#         if attendance_sheet.cell(row=row, column=2).value and attendance_sheet.cell(row=row, column=2).value.lower() == name.lower():
#             return row
#     return None

# def add_student_to_excel(roll_no, name):
#     new_row = attendance_sheet.max_row + 1
#     attendance_sheet.cell(row=new_row, column=1, value=roll_no)
#     attendance_sheet.cell(row=new_row, column=2, value=name)
#     return new_row

# # Load student data
# known_face_encodings = []
# known_face_names = []
# roll_number_mapping = {}

# def load_student_data():
#     global known_face_encodings, known_face_names, roll_number_mapping
    
#     try:
#         image_files = os.listdir(STUDENT_IMAGES_PATH)
#         print(f"Found {len(image_files)} student images")
#     except FileNotFoundError:
#         print(f"Error: Directory '{STUDENT_IMAGES_PATH}' not found.")
#         return
    
#     for img_name in image_files:
#         img_path = os.path.join(STUDENT_IMAGES_PATH, img_name)
#         img = cv2.imread(img_path)
#         if img is not None:
#             rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
#             encodings = face_recognition.face_encodings(rgb_img)
#             if encodings:
#                 known_face_encodings.append(encodings[0])
#                 known_face_names.append(os.path.splitext(img_name)[0])
    
#     try:
#         with open(ROLL_CSV_FILE, 'r') as f:
#             for line in f:
#                 parts = line.strip().split(',')
#                 if len(parts) >= 2:
#                     name = parts[0].strip().lower()
#                     roll_no = parts[1].strip()
#                     roll_number_mapping[name] = roll_no
#         print(f"Loaded {len(roll_number_mapping)} student mappings")
#     except FileNotFoundError:
#         print(f"Error: File '{ROLL_CSV_FILE}' not found.")

# load_student_data()

# @app.route('/')
# def index():
#     return render_template('index.html')

# @app.route('/api/students')
# def get_students():
#     students = []
#     for name in known_face_names:
#         roll_no = roll_number_mapping.get(name.lower(), 'N/A')
#         students.append({
#             'name': name,
#             'roll_no': roll_no
#         })
#     return jsonify({'students': students})

# @app.route('/api/detect', methods=['POST'])
# def detect_faces():
#     if 'image' not in request.files:
#         return jsonify({'error': 'No image provided'}), 400
    
#     file = request.files['image']
#     if file.filename == '':
#         return jsonify({'error': 'No selected file'}), 400
    
#     img_bytes = file.read()
#     img_array = np.frombuffer(img_bytes, np.uint8)
#     img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    
#     if img is None:
#         return jsonify({'error': 'Could not decode image'}), 400
    
#     rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
#     face_locations = face_recognition.face_locations(rgb_img)
#     face_encodings = face_recognition.face_encodings(rgb_img, face_locations)
    
#     if not face_encodings:
#         return jsonify({'detected': False})
    
#     for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
#         matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
#         face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
#         best_match_idx = np.argmin(face_distances)
        
#         if matches[best_match_idx] and face_distances[best_match_idx] < 0.6:
#             name = known_face_names[best_match_idx]
#             roll_no = roll_number_mapping.get(name.lower(), 'N/A')
            
#             height, width = img.shape[:2]
#             face_location = {
#                 'top': top / height,
#                 'right': right / width,
#                 'bottom': bottom / height,
#                 'left': left / width
#             }
            
#             return jsonify({
#                 'detected': True,
#                 'name': name,
#                 'roll_no': roll_no,
#                 'face_location': face_location
#             })
    
#     return jsonify({'detected': False})

# @app.route('/api/mark', methods=['POST'])
# def mark_attendance():
#     data = request.get_json()
#     roll_no = data.get('roll_no')
#     name = data.get('name')
    
#     if not roll_no or not name:
#         return jsonify({'success': False, 'error': 'Missing data'}), 400
    
#     try:
#         row = find_student_row(name)
#         if row is None:
#             row = add_student_to_excel(roll_no, name)
        
#         status_cell = attendance_sheet.cell(row=row, column=date_column)
#         if status_cell.value == 'Present':
#             return jsonify({
#                 'success': False,
#                 'message': f'{name} already marked present today'
#             })
        
#         timestamp = datetime.now().strftime('%H:%M:%S')
#         attendance_sheet.cell(row=row, column=date_column, value="Present")
#         attendance_sheet.cell(row=row, column=date_column+1, value=timestamp)
#         attendance_workbook.save(ATTENDANCE_EXCEL_FILE)
        
#         print(f"Marked attendance for {name} (Roll No: {roll_no}) at {timestamp}")
        
#         return jsonify({
#             'success': True,
#             'message': f'Attendance marked for {name}',
#             'timestamp': timestamp
#         })
#     except Exception as e:
#         print(f"Error marking attendance: {str(e)}")
#         return jsonify({
#             'success': False,
#             'error': str(e)
#         }), 500

# @app.route('/quit')
# def quit_app():
#     print("Application quit requested")
#     return "You can now close this window", 200

# if __name__ == '__main__':
#     app.run(debug=True)


from flask import Flask, render_template, request, jsonify, session
import cv2
import numpy as np
import os
import random
from datetime import datetime
import face_recognition
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Configuration
STUDENT_IMAGES_PATH = 'StudentImages'
ROLL_CSV_FILE = 'attendancenew.csv'
ATTENDANCE_EXCEL_FILE = 'attendancenew.xlsx'

# Initialize Excel file
def init_attendance_excel():
    if os.path.exists(ATTENDANCE_EXCEL_FILE):
        workbook = load_workbook(ATTENDANCE_EXCEL_FILE)
        sheet = workbook.active
        
        if sheet.cell(row=1, column=1).value != "Roll No.":
            sheet.cell(row=1, column=1, value="Roll No.")
            sheet.cell(row=1, column=2, value="Name")
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="Roll No.")
        sheet.cell(row=1, column=2, value="Name")
    
    return workbook, sheet

attendance_workbook, attendance_sheet = init_attendance_excel()
current_date = datetime.now().strftime('%Y-%m-%d')

def add_date_columns_if_needed():
    global attendance_sheet, current_date
    
    date_col = 3
    date_exists = False
    
    while True:
        header = attendance_sheet.cell(row=1, column=date_col).value
        if header is None:
            break
        if header == current_date:
            date_exists = True
            break
        date_col += 4
    
    if not date_exists:
        attendance_sheet.cell(row=1, column=date_col, value=current_date)
        attendance_sheet.cell(row=2, column=date_col, value="Status")
        attendance_sheet.cell(row=2, column=date_col+1, value="Timestamp")
    
    return date_col

date_column = add_date_columns_if_needed()

def find_student_row(name):
    for row in range(3, attendance_sheet.max_row + 1):
        if attendance_sheet.cell(row=row, column=2).value and attendance_sheet.cell(row=row, column=2).value.lower() == name.lower():
            return row
    return None

def add_student_to_excel(roll_no, name):
    new_row = attendance_sheet.max_row + 1
    attendance_sheet.cell(row=new_row, column=1, value=roll_no)
    attendance_sheet.cell(row=new_row, column=2, value=name)
    return new_row

# Load student data
known_face_encodings = []
known_face_names = []
roll_number_mapping = {}

def load_student_data():
    global known_face_encodings, known_face_names, roll_number_mapping
    
    try:
        image_files = os.listdir(STUDENT_IMAGES_PATH)
        print(f"Found {len(image_files)} student images")
    except FileNotFoundError:
        print(f"Error: Directory '{STUDENT_IMAGES_PATH}' not found.")
        return
    
    for img_name in image_files:
        img_path = os.path.join(STUDENT_IMAGES_PATH, img_name)
        img = cv2.imread(img_path)
        if img is not None:
            rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encodings = face_recognition.face_encodings(rgb_img)
            if encodings:
                known_face_encodings.append(encodings[0])
                known_face_names.append(os.path.splitext(img_name)[0])
    
    try:
        with open(ROLL_CSV_FILE, 'r') as f:
            for line in f:
                parts = line.strip().split(',')
                if len(parts) >= 2:
                    name = parts[0].strip().lower()
                    roll_no = parts[1].strip()
                    roll_number_mapping[name] = roll_no
        print(f"Loaded {len(roll_number_mapping)} student mappings")
    except FileNotFoundError:
        print(f"Error: File '{ROLL_CSV_FILE}' not found.")

load_student_data()

# Load Haar cascade for eye detection
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_eye.xml')

@app.route('/api/verify_eyes', methods=['POST'])
def verify_eye_movement():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    
    file = request.files['image']
    img_bytes = file.read()
    img_array = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    
    if img is None:
        return jsonify({'error': 'Could not decode image'}), 400

    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    eyes = eye_cascade.detectMultiScale(gray, 1.1, 4)
    
    if len(eyes) < 2:
        return jsonify({'detected': False, 'message': 'Eyes not detected'})
    
    # Store eye positions in session
    if 'eye_positions' not in session:
        session['eye_positions'] = []
    
    eye_centers = [(x + w//2, y + h//2) for (x,y,w,h) in eyes[:2]]
    session['eye_positions'].append(eye_centers)
    
    # Keep only last 5 frames
    if len(session['eye_positions']) > 5:
        session['eye_positions'] = session['eye_positions'][-5:]
    
    # Check for movement
    if len(session['eye_positions']) >= 3:
        movement = calculate_eye_movement(session['eye_positions'])
        if movement > 5:  # Threshold in pixels
            session.pop('eye_positions', None)  # Reset for next check
            return jsonify({
                'detected': True,
                'is_real': True,
                'movement': movement
            })
    
    return jsonify({
        'detected': True,
        'is_real': False,
        'message': 'Move your eyes side to side'
    })

def calculate_eye_movement(positions):
    """Calculate total eye movement variance"""
    movements = []
    for i in range(1, len(positions)):
        # Calculate distance between eye centers
        dx = positions[i][0][0] - positions[i-1][0][0]
        dy = positions[i][0][1] - positions[i-1][0][1]
        movements.append(np.sqrt(dx**2 + dy**2))
    return np.mean(movements)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/students')
def get_students():
    students = []
    for name in known_face_names:
        roll_no = roll_number_mapping.get(name.lower(), 'N/A')
        students.append({
            'name': name,
            'roll_no': roll_no
        })
    return jsonify({'students': students})

@app.route('/api/detect', methods=['POST'])
def detect_faces():
    if 'image' not in request.files:
        return jsonify({'error': 'No image provided'}), 400
    
    file = request.files['image']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    img_bytes = file.read()
    img_array = np.frombuffer(img_bytes, np.uint8)
    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
    
    if img is None:
        return jsonify({'error': 'Could not decode image'}), 400
    
    rgb_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_img)
    face_encodings = face_recognition.face_encodings(rgb_img, face_locations)
    
    if not face_encodings:
        return jsonify({'detected': False})
    
    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
        best_match_idx = np.argmin(face_distances)
        
        if matches[best_match_idx] and face_distances[best_match_idx] < 0.6:
            name = known_face_names[best_match_idx]
            roll_no = roll_number_mapping.get(name.lower(), 'N/A')
            
            height, width = img.shape[:2]
            face_location = {
                'top': top / height,
                'right': right / width,
                'bottom': bottom / height,
                'left': left / width
            }
            
            return jsonify({
                'detected': True,
                'name': name,
                'roll_no': roll_no,
                'face_location': face_location
            })
    
    return jsonify({'detected': False})

@app.route('/api/mark', methods=['POST'])
def mark_attendance():
    data = request.get_json()
    roll_no = data.get('roll_no')
    name = data.get('name')
    
    if not roll_no or not name:
        return jsonify({'success': False, 'error': 'Missing data'}), 400
    
    try:
        row = find_student_row(name)
        if row is None:
            row = add_student_to_excel(roll_no, name)
        
        status_cell = attendance_sheet.cell(row=row, column=date_column)
        if status_cell.value == 'Present':
            return jsonify({
                'success': False,
                'message': f'{name} already marked present today'
            })
        
        timestamp = datetime.now().strftime('%H:%M:%S')
        attendance_sheet.cell(row=row, column=date_column, value="Present")
        attendance_sheet.cell(row=row, column=date_column+1, value=timestamp)
        attendance_workbook.save(ATTENDANCE_EXCEL_FILE)
        
        print(f"Marked attendance for {name} (Roll No: {roll_no}) at {timestamp}")
        
        return jsonify({
            'success': True,
            'message': f'Attendance marked for {name}',
            'timestamp': timestamp
        })
    except Exception as e:
        print(f"Error marking attendance: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/quit')
def quit_app():
    print("Application quit requested")
    return "You can now close this window", 200


if __name__ == '__main__':
    app.run(debug=True)